import os
import secrets
import shutil
import zipfile
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook


ALLOWED_EXTS = (".png", ".jpg", ".jpeg")
RADIO_DICT = {1: "胃底", 2: "胃体", 3: "胃窦", 4: "胃角", 5: "幽门前区"}
APP_TITLE = "胃镜影像标注工作站"
SOURCE_DIR_NAME = "原始图片"
SELECTED_DIR_NAME = "拟入选"
DELETED_DIR_NAME = "拟删除"
EXCEL_NAME = "图片信息.xlsx"

NEW_STRENGTH_WEAK_BASES = [
    "黏膜肿胀",
    "点状发红",
    "鸡皮样改变",
    "白色浑浊黏液",
    "增生性息肉",
    "黄色瘤",
    "地图状发红",
    "藤壶征",
    "RAC",
    "划痕征",
    "胃底腺息肉",
    "陈旧性出血斑",
    "脊状发红",
    "隆起性弥漫",
    "多发白色扁平隆起",
    "斑状发红",
    "凹陷性糜烂",
    "胃体糜烂",
]
CORE_STRENGTH_WEAK_PAIRS = [
    ("萎缩", "萎缩（弱）"),
    ("肠化生", "肠化生（弱）"),
    ("弥漫性发红", "弥漫性发红（弱）"),
]
ALL_STRENGTH_WEAK_PAIRS = CORE_STRENGTH_WEAK_PAIRS + [
    ("皱襞肿大", "皱襞肿大（弱）"),
    ("结节性胃炎", "结节性胃炎（弱）"),
] + [(b, f"{b}（弱）") for b in NEW_STRENGTH_WEAK_BASES]

SYMPTOM_MUTEX_PAIRS = list(ALL_STRENGTH_WEAK_PAIRS)

SYMPTOM_LABELS: List[str] = []
for strong, weak in ALL_STRENGTH_WEAK_PAIRS:
    SYMPTOM_LABELS.extend([strong, weak])

EXCEL_FEATURE_LABELS = (
    ["萎缩", "肠化生", "弥漫性发红", "皱襞肿大", "结节性胃炎"] + NEW_STRENGTH_WEAK_BASES
)


def normalize_relative_path(raw_path: str) -> str:
    normalized = raw_path.replace("\\", "/").strip("/")
    cleaned_parts: List[str] = []
    for part in normalized.split("/"):
        p = part.strip()
        if not p or p in {".", ".."}:
            continue
        cleaned_parts.append(p)
    return "/".join(cleaned_parts)


def rel_to_parts(rel_path: str) -> List[str]:
    return [p for p in rel_path.split("/") if p]


def join_rel(base: str, rel_path: str) -> str:
    parts = rel_to_parts(rel_path)
    return os.path.join(base, *parts) if parts else base


def source_dir(workspace_dir: str) -> str:
    return os.path.join(workspace_dir, SOURCE_DIR_NAME)


def image_source_path(workspace_dir: str, image_name: str) -> str:
    return join_rel(source_dir(workspace_dir), image_name)


def selection_path(workspace_dir: str, image_name: str) -> str:
    return join_rel(os.path.join(workspace_dir, SELECTED_DIR_NAME), image_name)


def deletion_path(workspace_dir: str, image_name: str) -> str:
    return join_rel(os.path.join(workspace_dir, DELETED_DIR_NAME), image_name)


def list_images(folder_path: str) -> List[str]:
    out: List[str] = []
    for root, _, files in os.walk(folder_path):
        for name in files:
            if not name.lower().endswith(ALLOWED_EXTS):
                continue
            abs_path = os.path.join(root, name)
            rel_path = os.path.relpath(abs_path, folder_path).replace(os.sep, "/")
            out.append(rel_path)
    return sorted(out)


def ensure_excel(folder_path: str):
    excel_path = os.path.join(folder_path, EXCEL_NAME)
    expected_headers = ["序号", "图片名", "部位"] + EXCEL_FEATURE_LABELS

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        current_headers = [cell.value for cell in ws[1]]
        if current_headers != expected_headers:
            backup_path = os.path.join(folder_path, "图片信息_旧表头备份.xlsx")
            try:
                shutil.copy2(excel_path, backup_path)
            except Exception:
                pass
            wb = Workbook()
            ws = wb.active
            ws.append(expected_headers)
            wb.save(excel_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(expected_headers)
        wb.save(excel_path)

    return excel_path


def feature_values_from_symptoms(symptoms: List[str]) -> Dict[str, int]:
    s = set(symptoms)
    out: Dict[str, int] = {}
    for strong, weak in ALL_STRENGTH_WEAK_PAIRS:
        out[strong] = 2 if strong in s else (1 if weak in s else 0)
    return out


def status_for_image(folder_path: str, image_name: str) -> str:
    selected_path = selection_path(folder_path, image_name)
    deleted_path = deletion_path(folder_path, image_name)
    if os.path.exists(selected_path):
        return "拟入选"
    if os.path.exists(deleted_path):
        return "拟删除"
    return "待处理"


def check_all_processed(folder_path: str, image_files: List[str]) -> bool:
    for image_name in image_files:
        selected_path = selection_path(folder_path, image_name)
        deleted_path = deletion_path(folder_path, image_name)
        if not os.path.exists(selected_path) and not os.path.exists(deleted_path):
            return False
    return True


def remove_excel_rows_for_image(folder_path: str, image_name: str):
    excel_path = os.path.join(folder_path, EXCEL_NAME)
    if not os.path.exists(excel_path):
        return
    wb = load_workbook(excel_path)
    ws = wb.active
    for row in reversed(range(1, ws.max_row + 1)):
        if ws.cell(row=row, column=2).value == image_name:
            ws.delete_rows(row)
    wb.save(excel_path)


@dataclass
class SessionState:
    workspace_dir: Optional[str] = None
    folder_label: Optional[str] = None
    image_files: List[str] = field(default_factory=list)
    current_index: int = 0
    message: Optional[str] = None
    message_type: str = "info"


app = FastAPI()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SESSION_ROOT = os.path.join(BASE_DIR, "session_data")
os.makedirs(SESSION_ROOT, exist_ok=True)
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")

_SESSIONS: Dict[str, SessionState] = {}


def get_session_id(request: Request) -> str:
    sid = request.cookies.get("sid")
    if sid and sid in _SESSIONS:
        return sid
    sid = secrets.token_urlsafe(16)
    _SESSIONS[sid] = SessionState()
    return sid


def get_state(request: Request) -> SessionState:
    sid = get_session_id(request)
    return _SESSIONS[sid]


def set_message(state: SessionState, text: str, level: str = "info"):
    state.message = text
    state.message_type = level


def clear_state_data(state: SessionState):
    state.workspace_dir = None
    state.folder_label = None
    state.image_files = []
    state.current_index = 0


def wipe_workspace(workspace_dir: Optional[str]):
    if workspace_dir and os.path.isdir(workspace_dir):
        shutil.rmtree(workspace_dir, ignore_errors=True)


def ensure_workspace_for_session(sid: str) -> str:
    workspace_dir = os.path.join(SESSION_ROOT, sid)
    wipe_workspace(workspace_dir)
    os.makedirs(source_dir(workspace_dir), exist_ok=True)
    return workspace_dir


def infer_folder_label(paths: List[str]) -> Optional[str]:
    parts_list = [rel_to_parts(p) for p in paths if p]
    if not parts_list:
        return None
    if all(len(parts) > 1 for parts in parts_list):
        candidate = parts_list[0][0]
        if all(parts[0] == candidate for parts in parts_list):
            return candidate
    return None


def strip_common_root(rel_path: str, root_name: Optional[str]) -> str:
    parts = rel_to_parts(rel_path)
    if root_name and len(parts) > 1 and parts[0] == root_name:
        return "/".join(parts[1:])
    return "/".join(parts)


def build_result_zip(workspace_dir: str) -> str:
    zip_path = os.path.join(workspace_dir, "标注结果.zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        excel_path = os.path.join(workspace_dir, EXCEL_NAME)
        if os.path.exists(excel_path):
            zf.write(excel_path, arcname=EXCEL_NAME)
        for folder_name in (SELECTED_DIR_NAME, DELETED_DIR_NAME):
            folder_abs = os.path.join(workspace_dir, folder_name)
            if not os.path.isdir(folder_abs):
                continue
            for root, _, files in os.walk(folder_abs):
                for name in files:
                    abs_path = os.path.join(root, name)
                    rel_path = os.path.relpath(abs_path, workspace_dir)
                    zf.write(abs_path, arcname=rel_path)
    return zip_path


def _redirect(url: str, sid: str) -> RedirectResponse:
    resp = RedirectResponse(url=url, status_code=303)
    resp.set_cookie("sid", sid, httponly=True, samesite="lax")
    return resp


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]

    current_image = None
    statuses = []
    if state.workspace_dir and state.image_files:
        current_image = state.image_files[state.current_index]
        statuses = [
            {"name": name, "status": status_for_image(state.workspace_dir, name)}
            for name in state.image_files
        ]
    selected_count = sum(1 for item in statuses if item["status"] == "拟入选")
    deleted_count = sum(1 for item in statuses if item["status"] == "拟删除")
    pending_count = sum(1 for item in statuses if item["status"] == "待处理")

    ctx = {
        "request": request,
        "app_title": APP_TITLE,
        "folder_label": state.folder_label or "未加载文件夹",
        "has_folder": bool(state.workspace_dir),
        "image_count": len(state.image_files),
        "current_index": state.current_index,
        "current_image": current_image,
        "radio_dict": RADIO_DICT,
        "symptom_labels": SYMPTOM_LABELS,
        "symptom_mutex_pairs": SYMPTOM_MUTEX_PAIRS,
        "statuses": statuses,
        "message": state.message,
        "message_type": state.message_type,
        "selected_count": selected_count,
        "deleted_count": deleted_count,
        "pending_count": pending_count,
        "all_processed": bool(state.workspace_dir and check_all_processed(state.workspace_dir, state.image_files)),
    }
    state.message = None
    state.message_type = "info"
    resp = templates.TemplateResponse(request=request, name="index.html", context=ctx)
    resp.set_cookie("sid", sid, httponly=True, samesite="lax")
    return resp


@app.post("/set-folder")
def set_folder_legacy(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    set_message(state, "当前版本已改为网页文件夹选择，请点击“选择影像文件夹”。", "warning")
    return _redirect("/", sid)


@app.post("/upload-folder")
async def upload_folder(request: Request, files: List[UploadFile] = File(default=[])):
    sid = get_session_id(request)
    state = _SESSIONS[sid]

    if not files:
        set_message(state, "请选择一个包含图片的文件夹后再加载。", "warning")
        return _redirect("/", sid)

    raw_names: List[str] = []
    parsed_files: List[Tuple[UploadFile, str]] = []
    for upload in files:
        rel_path = normalize_relative_path(upload.filename or "")
        if not rel_path:
            await upload.close()
            continue
        raw_names.append(rel_path)
        parsed_files.append((upload, rel_path))

    folder_label = infer_folder_label(raw_names)
    workspace_dir = ensure_workspace_for_session(sid)
    source_root = source_dir(workspace_dir)

    saved_count = 0
    try:
        for upload, raw_rel in parsed_files:
            rel_path = strip_common_root(raw_rel, folder_label)
            if not rel_path.lower().endswith(ALLOWED_EXTS):
                await upload.close()
                continue
            dest_path = join_rel(source_root, rel_path)
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            with open(dest_path, "wb") as out_file:
                shutil.copyfileobj(upload.file, out_file)
            saved_count += 1
            await upload.close()
    finally:
        for upload in files:
            try:
                await upload.close()
            except Exception:
                pass

    image_files = list_images(source_root)
    if saved_count == 0 or not image_files:
        wipe_workspace(workspace_dir)
        clear_state_data(state)
        set_message(state, "未检测到可用图片，请选择 png/jpg/jpeg 文件。", "error")
        return _redirect("/", sid)

    ensure_excel(workspace_dir)
    state.workspace_dir = workspace_dir
    state.folder_label = folder_label or "已上传影像"
    state.image_files = image_files
    state.current_index = 0
    set_message(state, f"已加载 {len(image_files)} 张影像，可以开始标注。", "success")
    return _redirect("/", sid)


@app.post("/reset-session")
def reset_session(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    wipe_workspace(state.workspace_dir)
    clear_state_data(state)
    set_message(state, "已清空当前会话，请重新选择影像文件夹。", "info")
    return _redirect("/", sid)


@app.get("/image/current")
def get_current_image(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir or not state.image_files:
        raise HTTPException(status_code=404, detail="No image loaded")
    image_name = state.image_files[state.current_index]
    image_path = image_source_path(state.workspace_dir, image_name)
    if not os.path.exists(image_path):
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(image_path)


@app.post("/jump")
def jump(request: Request, index: int = Form(...)):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.image_files:
        return _redirect("/", sid)
    if 0 <= index < len(state.image_files):
        state.current_index = index
    return _redirect("/", sid)


@app.post("/previous")
def previous(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir or not state.image_files:
        return _redirect("/", sid)
    if state.current_index <= 0:
        set_message(state, "已经是第一张影像，无法继续回退。", "warning")
        return _redirect("/", sid)

    state.current_index -= 1
    image_name = state.image_files[state.current_index]

    selected_path = selection_path(state.workspace_dir, image_name)
    deleted_path = deletion_path(state.workspace_dir, image_name)
    if os.path.exists(selected_path):
        os.remove(selected_path)
    if os.path.exists(deleted_path):
        os.remove(deleted_path)
    remove_excel_rows_for_image(state.workspace_dir, image_name)

    set_message(state, "已回到上一张，并撤销该张的标注结果。", "info")
    return _redirect("/", sid)


@app.post("/delete")
def delete_current(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir or not state.image_files:
        return _redirect("/", sid)

    image_name = state.image_files[state.current_index]
    deleted_folder = os.path.join(state.workspace_dir, DELETED_DIR_NAME)
    os.makedirs(deleted_folder, exist_ok=True)
    source_path = image_source_path(state.workspace_dir, image_name)
    dest_path = deletion_path(state.workspace_dir, image_name)
    selected_path = selection_path(state.workspace_dir, image_name)
    if os.path.exists(selected_path):
        os.remove(selected_path)
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    shutil.copy2(source_path, dest_path)

    remove_excel_rows_for_image(state.workspace_dir, image_name)

    state.current_index += 1
    if state.current_index >= len(state.image_files):
        if check_all_processed(state.workspace_dir, state.image_files):
            state.current_index = len(state.image_files) - 1
            set_message(state, "所有影像已处理完毕。", "success")
        else:
            state.current_index = len(state.image_files) - 1
            set_message(state, "请继续完成剩余影像。", "info")
    else:
        set_message(state, "当前影像已移出批次。", "warning")
    return _redirect("/", sid)


@app.post("/confirm")
def confirm_current(
    request: Request,
    part: Optional[int] = Form(default=None),
    symptoms: List[str] = Form(default=[]),
):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir or not state.image_files:
        return _redirect("/", sid)

    radio_result = RADIO_DICT.get(part, "")
    features = feature_values_from_symptoms(symptoms)
    check_results = [features[label] for label in EXCEL_FEATURE_LABELS]

    if not radio_result or not any(v > 0 for v in check_results):
        set_message(state, "请至少选择 1 个部位和 1 个症状后再保存。", "error")
        return _redirect("/", sid)

    selected_folder = os.path.join(state.workspace_dir, SELECTED_DIR_NAME)
    os.makedirs(selected_folder, exist_ok=True)
    image_name = state.image_files[state.current_index]
    source_path = image_source_path(state.workspace_dir, image_name)
    dest_path = selection_path(state.workspace_dir, image_name)
    deleted_path = deletion_path(state.workspace_dir, image_name)
    if os.path.exists(deleted_path):
        os.remove(deleted_path)
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    shutil.copy2(source_path, dest_path)

    remove_excel_rows_for_image(state.workspace_dir, image_name)
    excel_path = ensure_excel(state.workspace_dir)
    wb = load_workbook(excel_path)
    ws = wb.active
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=next_row - 1)
    ws.cell(row=next_row, column=2, value=image_name)
    ws.cell(row=next_row, column=3, value=radio_result)
    for i, result in enumerate(check_results):
        ws.cell(row=next_row, column=4 + i, value=result)
    wb.save(excel_path)

    state.current_index += 1
    if state.current_index >= len(state.image_files):
        if check_all_processed(state.workspace_dir, state.image_files):
            state.current_index = len(state.image_files) - 1
            set_message(state, "所有影像已处理完毕。", "success")
        else:
            state.current_index = len(state.image_files) - 1
            set_message(state, "请继续完成剩余影像。", "info")
    else:
        set_message(state, "已保存当前标注，自动进入下一张。", "success")
    return _redirect("/", sid)


@app.get("/download/excel")
def download_excel(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir:
        raise HTTPException(status_code=404, detail="No session workspace")
    excel_path = os.path.join(state.workspace_dir, EXCEL_NAME)
    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Excel not ready")
    filename = f"{state.folder_label or '胃镜影像'}_图片信息.xlsx"
    return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)


@app.get("/download/package")
def download_package(request: Request):
    sid = get_session_id(request)
    state = _SESSIONS[sid]
    if not state.workspace_dir:
        raise HTTPException(status_code=404, detail="No session workspace")
    zip_path = build_result_zip(state.workspace_dir)
    if not os.path.exists(zip_path):
        raise HTTPException(status_code=404, detail="Package not ready")
    filename = f"{state.folder_label or '胃镜影像'}_标注结果.zip"
    return FileResponse(zip_path, media_type="application/zip", filename=filename)


@app.get("/health")
def health():
    return {"ok": True}

